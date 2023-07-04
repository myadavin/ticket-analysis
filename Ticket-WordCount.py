import xlrd
from collections import Counter
from openpyxl import Workbook

# Read the XLS file
#workbook = xlrd.open_workbook('your_file.xls')
workbook = xlrd.open_workbook('/Users/myadav/Documents/myadav/home/VMWare/Horizon/My_Files/Download/xls/cdto-map-ticket-data.xls')
sheet = workbook.sheet_by_index(0)  # Assuming data is in the first sheet

# Extract text from cells and concatenate into a single string
text = ""
for row in range(sheet.nrows):
    for col in range(sheet.ncols):
        cell_value = sheet.cell_value(row, col)
        if isinstance(cell_value, str):
            text += cell_value + " "

# Split the text into words and count occurrences
word_count = Counter(text.split())

# Create a new workbook and select the active sheet
output_workbook = Workbook()
output_sheet = output_workbook.active

# Write the word count to the XLS file
output_sheet.cell(row=1, column=1, value="Word")
output_sheet.cell(row=1, column=2, value="Count")

row_index = 2
for idx, (word, count) in enumerate(word_count.most_common(), start=1):
    output_sheet.cell(row=row_index, column=1, value=word)
    output_sheet.cell(row=row_index, column=2, value=count)
    row_index += 1

# Save the XLS file
output_workbook.save('word_count_output.xls')

print("Word count output written to word_count_output.xls")
